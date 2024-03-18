from pathlib import Path
import pandas as pd
import openpyxl
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.cell.cell import Cell

from app.ErrorHandler import CustomError


class ExcelTemplateProcessor:
    """excelテンプレートファイルを扱うクラス"""

    def __init__(self, template_file_path: Path) -> None:
        self.wb = self._load_workbook(template_file_path)

    def _load_workbook(self, template_file_path: Path) -> Workbook:
        try:
            wb = openpyxl.load_workbook(template_file_path)
            return wb
        except FileNotFoundError as e:
            raise CustomError(
                f"can not find specified file path:{template_file_path}"
            ) from e  #  from e を付けておくことで上位で元のエラーを表示できる。e.__cause__

    def dataframe_to_excel(self, df: pd.DataFrame, keyword: str) -> None:
        """指定した単語が存在するセルを起点としてdfを貼り付ける"""
        cell = self._search_keyword_in_workbook(keyword)
        ws: Worksheet = cell.parent
        start_row = cell.row
        start_col = cell.column
        for r_idx, row in enumerate(df.values, start=start_row):
            for c_idx, value in enumerate(row, start=start_col):
                ws.cell(row=r_idx, column=c_idx).value = value

    def _search_keyword_in_workbook(self, keyword: str) -> Cell:
        """指定された単語が含まれるcell object"""
        for ws in self.wb:
            for row in ws.rows:
                for cell in row:
                    if str(cell.value) == keyword:
                        return cell
        raise CustomError(f"can not find keyword '{keyword}' in workbook")

    # def save_as_excel(self):
    #     timestamp = datetime.now().strftime("%Y%m%d%H%M%S")
    #     self.wb.save(f"output_{timestamp}.xlsx")
